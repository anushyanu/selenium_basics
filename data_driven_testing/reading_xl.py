import openpyxl #this only works for xlsx extension
#excel is stored in these
# file -> workbook ->sheet -> rows ->columns

file ="/home/anushya/Documents/selenium.xlsx"  # this is absolute path
workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"]
rows = sheet.max_row
col = sheet.max_column

for r in range(1,rows+1):
    for c in range(1,col+1):
        print(sheet.cell(r,c).value, end=" ")
    print()
